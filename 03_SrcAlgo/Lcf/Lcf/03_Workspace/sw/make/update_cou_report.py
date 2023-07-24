"""
Created on 19.04.2017

@author: uidv4529
"""

import argparse
import logging
import os
import sys
from shutil import copy2
from xml.etree import ElementTree
import json

try:
    import openpyxl as px
except:
    pass


def exit_program(msg, logg):
    """will exit with msg
    """
    logg.critical(msg)
    print msg
    sys.exit()


def get_source_files_json(jsonf):
    """search for the source code files (iut=True)
    """
    with open(jsonf) as data_file:
        data = json.load(data_file)

    m_active = []

    for m in data['test_frames']:
        if m['enabled']:
            m_active.append(str(m['short_name']))
    return m_active


def get_source_files(cou_root):
    """search for the source code files (exclude headers)
    """
    comp_xml = CompositeReader(cou_root)
    comp_xml.read_content()

    return comp_xml.compos_names


def get_rev_info(rev_info):
    """provides revision info for the sut
    """
    rev_mem = {}
    for k in rev_info.keys():
        rev_xml = CompositeReader(rev_info[k])
        rev_xml.read_content()
        rev_mem[k] = rev_xml.rev_info

    return rev_mem


def get_exec_status(cou_root, fpath):
    """will get execution and generated status
    and return tuple (src_name,executed,generated)
    """
    if fpath.endswith('.xml'):
        all_src = get_source_files(fpath)
    else:
        all_src = get_source_files_json(fpath)
    rev_info = {}
    res = []

    for cur_dir in all_src:
        p = os.path.join(cou_root, cur_dir)
        if os.path.exists(p):
            ex = True
            done_dir0 = os.path.join(p, 'DONE')
            done_dir1 = os.path.join(p, cur_dir + '_DONE')
            if os.path.exists(done_dir0):
                done_dir = done_dir0
            else:
                done_dir = done_dir1

            zip_file = os.path.join(done_dir, cur_dir + '.zip')
            if os.path.isfile(zip_file):
                gen = True
            else:
                gen = False

            revision_file = os.path.join(done_dir, 'revision_info.xml')
            if os.path.exists(revision_file):
                rev_info[cur_dir] = revision_file
        else:
            ex = False
            gen = False
        res.append((cur_dir, ex, gen))

    return res, rev_info


def get_common_data(fpaths, all_src, logg):
    """provides the data in one merged variable
    """
    try:
        assert (len(fpaths) <= len(all_src))
    except:
        msg = 'the productive source files should be more then resulted xml files!'
        exit_program(msg, logg)

    res = []
    for src in all_src:
        res.append((get_path(fpaths, src[0]),
                    src[0],
                    src[1],
                    src[2]))
    res.sort()
    return res


def get_path(fpaths, dir_name):
    """returns abs path to the xml file
    """
    res = ''
    for p in fpaths:
        if p[0] == dir_name:
            res = p[1]
    return res


def get_xml_files(cou_root):
    """search for all xml files in entire folder
    """
    xmls = []
    postfix = '-summary.xml'
    for root, _, files in os.walk(cou_root):
        for f in files:
            if f.endswith(postfix):
                dirn = f[:f.index(postfix)]
                xmls.append((dirn, os.path.join(root, f)))
    return xmls


def check_and_copy(rfile, logg):
    """ensure that target file fresh
    """
    if not os.path.exists(rfile):
        msg = 'the file (%s) is not exist, please provide correct path to the xlsx file' % rfile
        exit_program(msg, logg)

    fname = os.path.basename(rfile)
    dirname = os.path.dirname(rfile)
    name, ext = os.path.splitext(fname)

    if ext != '.xlsx':
        msg = 'the report file is not xlsx file'
        exit_program(msg, logg)

    nfile = name + '_new' + ext
    new_rfile = os.path.join(dirname, nfile)
    try:
        if os.path.exists(new_rfile):
            os.remove(new_rfile)
        copy2(rfile, new_rfile)

        return new_rfile

    except:
        msg = 'cannot remove the old file, please close the file (%s) first and try again' % new_rfile
        exit_program(msg, logg)


def display_percentage(sum_values, declt):
    return "{:^5.4}".format((sum_values / float(max(1, declt))) * 100.)


def create_txt_report_ctc(out_dir, cou_data, rev_info, a_logger):
    """will create txt file report for CTC Tool results
    """
    txt = []
    declc = 0
    declt = 0
    funclc = 0
    funclt = 0
    statlc = 0
    statlt = 0
    mcdclc = 0
    mcdclt = 0

    for cou in cou_data:
        declc += int(cou.cov_decision.lc)
        declt += int(cou.cov_decision.lt)
        funclc += int(cou.cov_function.lc)
        funclt += int(cou.cov_function.lt)
        statlc += int(cou.cov_statement.lc)
        statlt += int(cou.cov_statement.lt)
        mcdclc += int(cou.cov_mcdc.lc)
        mcdclt += int(cou.cov_mcdc.lt)

    dl = display_percentage(declc, declt)
    fl = display_percentage(funclc, funclt)
    sl = display_percentage(statlc, statlt)
    ml = display_percentage(mcdclc, mcdclt)

    rep = """
=========================================================================
=          Consolidated Coverage Report                                 =
=========================================================================
= Metrics       = Statements   = Decision   =  Functions   =  MC / DC   =
=========================================================================
= Total         =     %4d     =   %4d      =   %4d        =   %4d
= Executed      =     %4d     =   %4d      =   %4d        =   %4d
= Percentage(%%) =    %s     =   %s     =   %s       =   %s
=========================================================================

""" % (statlt, declt, funclt, mcdclt,
       statlc, declc, funclc, mcdclc,
       sl, dl, fl, ml)
    txt.append(rep)

    txt_format = \
        '{0:5s} | {1:40s} | {2:11s} | {3:13s} | {4:15s} | {5:15s} | {6:22s} | {7:22s} | {8:22s} | {9:22s} | {10:22s}\n'
    h = ['Sl No',
         'File Name',
         'is Executed',
         'Report avail.',
         'Passed TCs(No.)',
         'Failed TCs(No.)',
         'Statement Cov(%)',
         'Decision Cov(%)',
         'Function Cov(%)',
         'MC/DC Cov(%)',
         'Tested by']
    txt.append(txt_format.format(h[0], h[1], h[2], h[3], h[4], h[5], h[6], h[7], h[8], h[9], h[10]))
    for counter, cou in enumerate(cou_data):
        try:
            rev_mem = rev_info[cou.fname]
            fname = '%s (rev:%s)' % (cou.fname, rev_mem)
        except:
            fname = cou.fname
        txt.append(txt_format.format(str(counter + 1),
                                     fname,
                                     str(cou.executed),
                                     str(cou.generated),
                                     cou.passed,
                                     cou.failed,
                                     cou.cov_statement,
                                     cou.cov_decision,
                                     cou.cov_function,
                                     cou.cov_mcdc,
                                     cou.user))

    out_file = os.path.join(out_dir, 'cou_summary_ctc.txt')
    create_txt_file(out_file, txt, a_logger)


def create_txt_report_lcov(out_dir, cou_data, rev_info, a_logger):
    """will create txt file report for LCOV Tool results
    """
    txt = []
    llc = 0
    llt = 0
    blc = 0
    blt = 0
    flc = 0
    flt = 0
    for cou in cou_data:
        llc += int(cou.cov_line.lc)
        llt += int(cou.cov_line.lt)
        blc += int(cou.cov_branch.lc)
        blt += int(cou.cov_branch.lt)
        flc += int(cou.cov_function.lc)
        flt += int(cou.cov_function.lt)

    ll = display_percentage(llc, llt)
    bl = display_percentage(blc, blt)
    fl = display_percentage(flc, flt)

    rep = """
========================================================
=          Consolidated Coverage Report                =
========================================================
= Metrics       = Statements   = Branches   =  Functions
========================================================
= Total         =     %4d     =   %4d      =   %4d
= Executed      =     %4d     =   %4d      =   %4d
= Percentage(%%) =    %s     =   %s     =   %s
========================================================

""" % (llt, blt, flt, llc, blc, flc, ll, bl, fl)
    txt.append(rep)

    txt_format = '{0:5s} | {1:40s} | {2:11s} | {3:13s} | {4:20s} | {5:20s} | {6:20s} | {7:20s} | {8:22s} | {9:20s}\r'
    h = ['Sl No',
         'File Name',
         'is Executed',
         'Report avail..',
         'Passed TCs(No.)',
         'Failed TCs(No.)',
         'Line Coverage(%)',
         'Branch Coverage(%)',
         'Function Coverage(%)',
         'Tested by']
    txt.append(txt_format.format(h[0], h[1], h[2], h[3], h[4], h[5], h[6], h[7], h[8], h[9]))
    cnt = 1
    for cou in cou_data:
        try:
            rev_mem = rev_info[cou.fname]
            fname = '%s (rev:%s)' % (cou.fname, rev_mem)
        except:
            fname = cou.fname
        txt.append(txt_format.format(str(cnt),
                                     fname,
                                     str(cou.executed),
                                     str(cou.generated),
                                     cou.passed,
                                     cou.failed,
                                     cou.cov_line,
                                     cou.cov_branch,
                                     cou.cov_function,
                                     cou.user))
        cnt += 1
    out_file = os.path.join(out_dir, 'cou_summary_lcov.txt')
    create_txt_file(out_file, txt, a_logger)


def create_txt_file(out_file, txt_l, a_logger):
    """create the TXT file and store the results there.
    """
    with open(out_file, 'wb') as f:
        f.writelines(txt_l)

    msg = "created file %s " % os.path.basename(out_file)
    a_logger.info(msg)


def create_logger(log_fname_path):
    logger = logging.getLogger(__name__)
    handler = logging.FileHandler(log_fname_path)
    formatter = logging.Formatter(
        '%(asctime)s - [%(filename)2s:%(lineno)s - %(funcName)20s()] %(levelname)s -> %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    logger.setLevel(logging.DEBUG)
    return logger


def is_cov_tool_used(xml_list):
    """returns if the used COV tool
    """
    avail_res = [t[0] for t in xml_list if t[0] and t[3]]
    if avail_res:
        cou = CouXmlParser(avail_res[0])
        return cou.is_lcov_used

    return True


def add_data_lcov(xmlf):
    """will parse data and dump the data into class
    """
    if not (xmlf[0] and xmlf[3]):
        return CouDataLCOV(generated=xmlf[3],
                           executed=xmlf[2],
                           fname=xmlf[1])

    cou = CouXmlParser(xmlf[0])

    return CouDataLCOV(user=cou.user_name,
                       generated=xmlf[3],
                       executed=xmlf[2],
                       fname=cou.name,
                       passed=cou.passed,
                       failed=cou.failed,
                       cov_line=cou.cov_line,
                       cov_function=cou.cov_function,
                       cov_branch=cou.cov_branch)


def add_data_ctc(xmlf):
    """will parse data and dump the data into class
    """
    if not (xmlf[0] and xmlf[3]):
        return CouDataCTC(generated=xmlf[3],
                          executed=xmlf[2],
                          fname=xmlf[1])

    cou = CouXmlParser(xmlf[0])

    return CouDataCTC(cov_decision=cou.cov_decision,
                      cov_mcdc=cou.cov_mcdc,
                      cov_statement=cou.cov_statement,
                      cov_function=cou.cov_function,
                      user=cou.user_name,
                      generated=xmlf[3],
                      executed=xmlf[2],
                      fname=cou.name,
                      passed=cou.passed,
                      failed=cou.failed)


class CoverageResult(object):
    def __init__(self, cov_root=None):
        if cov_root is None:
            self.cov = '0'
            self.lc = '0'
            self.lt = '0'
        else:
            self.cov = cov_root.find('percentage').text
            self.lc = cov_root.find('hits').text
            self.lt = cov_root.find('total').text
        
    def __str__(self):
        return '%s%% (%s of %s)' % (self.cov, self.lc, self.lt)


class CoverageResultNull(CoverageResult):
    def __init__(self):
        CoverageResult.__init__(self)


def parse_xml(xml_list, is_lcov_used, a_logger):
    """this will parse xml data and store them into class
    """
    parsed_data = []
    for xml in xml_list:
        try:
            if is_lcov_used:
                cou_data = add_data_lcov(xml)
            else:
                cou_data = add_data_ctc(xml)

            parsed_data.append(cou_data)

        except:
            msg = 'Unexpected error happened while xml file (%s) parsing, keep parsing the next file' % xml
            a_logger.warning(msg)

    msg = 'in total %s xml files have been successfully parsed' % str(len(parsed_data))
    a_logger.info(msg)

    return parsed_data


class CompositeReader(object):
    """will read content of composite.xml file
    """
    namespace = 'http://contiautomotive.com/courage/coxml'

    def __init__(self, fpath):
        """init"""
        self.fpath = fpath
        self.root = None

    def read_content(self):
        """get content as string
        """
        with open(self.fpath, 'rt') as f:
            content = f.read().replace('\n', '')
            self.root = ElementTree.fromstring(content)
            self.remove_namespace()

    def remove_namespace(self):
        """Remove namespace in the passed document in place"""
        ns = u'{%s}' % self.namespace
        nsl = len(ns)
        for elem in self.root.getiterator():
            if elem.tag.startswith(ns):
                elem.tag = elem.tag[nsl:]

    @property
    def rev_info(self):
        """returns the revision info
        """
        return self.root.find("./project/files/pc/M/R").text

    @property
    def compos_names(self):
        """get all productive source file names
        """
        comps = self.root.find("./composites")
        comps_names = []
        for comp in comps.findall('composite'):
            comps_names.append(comp.attrib["name"])

        return comps_names


class CouXmlParser(object):
    """this will parse the summary xml file
    """
    namespace = 'http://contiautomotive.com/courage/coxml_result'

    def __init__(self, fpath):
        self.f = fpath
        self.root = None
        self.parse()

    def parse(self):
        """do parse
        """
        with open(self.f, 'rt') as f:
            content = f.read().replace('\n', '')

        self.root = ElementTree.fromstring(content)
        self.remove_namespace(self.root)

    def remove_namespace(self, doc):
        """Remove namespace in the passed document in place"""
        ns = u'{%s}' % self.namespace
        nsl = len(ns)
        for elem in doc.getiterator():
            if elem.tag.startswith(ns):
                elem.tag = elem.tag[nsl:]

    @property
    def cov_tool(self):
        return self.root.find("./TEST/code_coverages/tool").text

    @property
    def is_lcov_used(self):
        return True if self.cov_tool == 'LCOV' else False

    @property
    def user_name(self):
        return self.root.find("./username").text

    @property
    def failed(self):
        return self.root.find("./TEST/tests/failed").text

    @property
    def passed(self):
        return self.root.find("./TEST/tests/passed").text
    #
    # @property
    # def hostname(self):
    #     return self.root.find("./hostname").text

    # @property
    # def tool(self):
    #     return self.root.find("./tool").text
    #
    # @property
    # def project(self):
    #     return self.root.find("./project/project_id").text
    #
    # @property
    # def errors(self):
    #     return self.root.find("./log/errors").text
    #
    # @property
    # def infos(self):
    #     return self.root.find("./log/infos").text
    #
    # @property
    # def warnings(self):
    #     return self.root.find("./log/warnings").text
    #
    # @property
    # def finished(self):
    #     return self.root.find("./log/finished").text
    #
    # @property
    # def started(self):
    #     return self.root.find("./log/started").text

    @property
    def name(self):
        return self.root.find("./composite").attrib["name"]

    @property
    def cov_decision(self):
        return self.search_cov('decision')

    @property
    def cov_mcdc(self):
        return self.search_cov('mcdc')

    @property
    def cov_statement(self):
        return self.search_cov('statement')

    @property
    def cov_function(self):
        return self.search_cov('function')

    @property
    def cov_line(self):
        return self.search_cov('line')

    @property
    def cov_branch(self):
        return self.search_cov('branch')

    def search_cov(self, cov_type):
        """returns the coverage data
        """
        found_cov = None
        for cov in self.root.findall("./TEST/code_coverages/code_coverage"):
            if cov.attrib["type"] != cov_type:
                continue
            found_cov = cov
            break
        return CoverageResult(found_cov)

    # def get_asil(self):
    #     return self.root.find("./composite/alloc_sw_requirements/asil").text
    #
    # def get_test_f(self):
    #     return self.root.find("./TEST/tests/failed").text
    #
    # def get_test_p(self):
    #     return self.root.find("./TEST/tests/passed").text
    #
    # def get_test_s(self):
    #     return self.root.find("./TEST/tests/skipped").text
    #
    # def get_test_t(self):
    #     return self.root.find("./TEST/tests/total").text
    #
    # def get_test_perc(self):
    #     return self.root.find("./TEST/tests/percentage").text
    #
    # def get_test_assert_f(self):
    #     return self.root.find("./TEST/tests/assertions/failed").text
    #
    # def get_test_assert_p(self):
    #     return self.root.find("./TEST/tests/assertions/passed").text
    #
    # def get_test_assert_t(self):
    #     return self.root.find("./TEST/tests/assertions/total").text
    #
    # def get_tsuit_calls(self):
    #     return self.root.find("./TEST/testsuites").attrib["calls"]
    #
    # def get_tsuit_errs(self):
    #     return self.root.find("./TEST/testsuites").attrib["errors"]
    #
    # def get_tsuit_fails(self):
    #     return self.root.find("./TEST/testsuites").attrib["failures"]
    #
    # def get_tsuit_skipped(self):
    #     return self.root.find("./TEST/testsuites").attrib["skipped"]
    #
    # def get_tsuit_tests(self):
    #     return self.root.find("./TEST/testsuites").attrib["tests"]
    #
    # def get_tsuit_time(self):
    #     return self.root.find("./TEST/testsuites").attrib["time"]
    #
    # def get_tsuit_total(self):
    #     return len(self.root.findall("./TEST/testsuites"))


class CouData(object):
    """place to store the COU statistic data for each file
    """

    def __init__(self, generated, executed, fname, user='', passed='0', failed='0'):
        self.user = user
        self.generated = generated
        self.executed = executed
        self.fname = fname
        self.passed = passed
        self.failed = failed


class CouDataLCOV(CouData):
    """subclass from CouData to store the LCOV data
    """

    def __init__(self, cov_line=None, cov_function=None, cov_branch=None, *args, **kwargs):
        super(CouDataLCOV, self).__init__(*args, **kwargs)
        self.cov_line = cov_line if cov_line is not None else CoverageResultNull()
        self.cov_function = cov_function if cov_function is not None else CoverageResultNull()
        self.cov_branch = cov_branch if cov_branch is not None else CoverageResultNull()


class CouDataCTC(CouData):
    """subclass from CouData to store the CTC data
    """

    def __init__(self, cov_decision=None, cov_mcdc=None, cov_statement=None, cov_function=None, *args, **kwargs):
        super(CouDataCTC, self).__init__(*args, **kwargs)
        self.cov_decision = cov_decision if cov_decision is not None else CoverageResultNull()
        self.cov_mcdc = cov_mcdc if cov_mcdc is not None else CoverageResultNull()
        self.cov_statement = cov_statement if cov_statement is not None else CoverageResultNull()
        self.cov_function = cov_function if cov_function is not None else CoverageResultNull()


class UpdateCouReport(object):
    """will update provided xlsx file with parsed data from xml report summary
    """

    def __init__(self, fpath, cou_data, component_name, a_logger):
        self.fpath = fpath
        self.cou_data = cou_data
        self.component_name = component_name
        self.logger = a_logger

        self.logger.info(msg='Python openpyxl version....: %s\n' % px.__version__)

    def do_write_data(self):
        """ this will write data into xlsx file
        """
        wb = px.load_workbook(filename=self.fpath)
        self.logger.info(msg='document loaded OK %s' % self.fpath)

        ws = wb.get_sheet_by_name(name=self.component_name)
        header = self.get_ws_header(ws)
        position_idx = wb.get_index(ws)

        wb.remove_sheet(ws)

        ws_new = wb.create_sheet(title=self.component_name,
                                 index=position_idx)

        # write header
        for k in header:
            ws_new.cell(k).value = header[k]

        cnt = 2
        for cou in self.cou_data:
            self.write_rows(ws_new, cnt, cou)
            cnt += 1

        wb.save(filename=self.fpath)
        self.logger.info(msg="the report saved successfully")

        wb = px.load_workbook(filename=self.fpath)
        self.apply_styles(wb)
        wb.save(filename=self.fpath)
        self.logger.info(msg='styles are applied and saved')

    def apply_styles(self, wb):
        """set desired styles for the handover document
        """
        ws = wb.get_sheet_by_name(name=self.component_name)
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 20
        ws.column_dimensions["I"].width = 20
        ws.column_dimensions["J"].width = 20

    def write_rows(self, ws, cnt, cou):
        """this will write data line by line
        """
        ws.cell(coordinate=self.get_coord(cnt, 'A')).value = cnt - 1
        ws.cell(coordinate=self.get_coord(cnt, 'B')).value = cou.fname
        ws.cell(coordinate=self.get_coord(cnt, 'C')).value = self.get_executed(cou.executed)
        ws.cell(coordinate=self.get_coord(cnt, 'D')).value = self.get_generated(cou.generated)
        ws.cell(coordinate=self.get_coord(cnt, 'E')).value = cou.passed
        ws.cell(coordinate=self.get_coord(cnt, 'F')).value = cou.failed
        ws.cell(coordinate=self.get_coord(cnt, 'G')).value = cou.cov_line
        ws.cell(coordinate=self.get_coord(cnt, 'H')).value = cou.cov_function
        ws.cell(coordinate=self.get_coord(cnt, 'I')).value = cou.cov_branch
        ws.cell(coordinate=self.get_coord(cnt, 'J')).value = cou.user

    @staticmethod
    def get_generated(b_value):
        """based on the boolean value will return generated txt
        """

        return 'Generated' if b_value else 'Not generated'

    @staticmethod
    def get_executed(b_value):
        """based on the boolean value will return executed txt
        """
        return 'Executed' if b_value else 'Not executed'

    @staticmethod
    def get_ws_header(ws):
        """return the content of worksheet
        """
        d = {'A1': ws.cell('A1').value, 'B1': ws.cell('B1').value, 'C1': ws.cell('C1').value, 'D1': ws.cell('D1').value,
             'E1': ws.cell('E1').value, 'F1': ws.cell('F1').value, 'G1': ws.cell('G1').value, 'H1': ws.cell('H1').value,
             'I1': ws.cell('I1').value, 'J1': ws.cell('J1').value}

        return d

    @staticmethod
    def get_coord(row, column):
        """this will convert coordinate from (row=digit,column=letter) into A1 format
        """
        return column + str(row)


def analyze_parameters():
    arg_parser = argparse.ArgumentParser()
    arg_parser.add_argument("--cou_report_root", "-r", default="",
                            help="MANDATORY: abs path to the root folder of courage reports")
    arg_parser.add_argument("--outdir", "-o", default="",
                            help="MANDATORY: abs path to the results")
    arg_parser.add_argument("--report_file", "-x", default="",
                            help="OPTIONAL: abs path to the xlsx file which should be updated")
    arg_parser.add_argument("--composite_xml_file", "-xml", default="",
                            help="OPTIONAL: abs path to the the composite.xml file")
    arg_parser.add_argument("--logFileName", '-l', default="courage_report.log",
                            help="OPTIONAL: Log file name")
    arg_parser.add_argument("--txtReport", '-t', action="store_true",
                            help="OPTIONAL: only provide TXT file summary")
    arg_parser.add_argument("--json_file", "-j", default="",
                            help="OPTIONAL: abs path to the json file")
    arg_parser.add_argument("--compName", '-c', default="",
                            help="OPTIONAL: The name of the working sheet (CompName)"
                                 " in the report xlsx file which should be updated")
    args = arg_parser.parse_args()
    return args


def main():
    args = analyze_parameters()

    # args.outdir = r'D:\Sandbox\Variant\EM\05_Testing\05_Test_Environment\algo\modtests\courage_tests'
    # args.cou_report_root=r'D:\Sandbox\Variant\EM\05_Testing\05_Test_Environment\algo\modtests\courage_tests\out\EM'
    # args.composite_xml_file =
    # r'D:\Sandbox\Variant\EM\05_Testing\05_Test_Environment\algo\modtests\courage_tests\composites.xml'
    # args.txtReport=True
    # args.json_file = r'C:\Users\uidv4529\Documents\SametimeFileTransfers\test_frames1.json'

    out_dir = args.outdir

    my_logger = create_logger(os.path.join(out_dir, args.logFileName))

    if args.composite_xml_file:
        must_run_mods = args.composite_xml_file
    elif args.json_file:
        must_run_mods = args.json_file
    else:
        exit_program('one of the optional inputs missing', my_logger)
        return

    stat, rev_info = get_exec_status(args.cou_report_root, must_run_mods)

    rev_info_parsed = get_rev_info(rev_info)

    xml_path = get_xml_files(args.cou_report_root)

    xml_files = get_common_data(xml_path, stat, my_logger)

    lcov_tool_used = is_cov_tool_used(xml_files)

    cou_data = parse_xml(xml_files, lcov_tool_used, my_logger)

    if not args.txtReport:
        copied_file = check_and_copy(args.report_file, my_logger)

        cou_report = UpdateCouReport(fpath=copied_file,
                                     cou_data=cou_data,
                                     component_name=args.compName,
                                     a_logger=my_logger)
        cou_report.do_write_data()

    else:
        if lcov_tool_used:
            create_txt_report_lcov(out_dir, cou_data, rev_info_parsed, my_logger)
        else:
            create_txt_report_ctc(out_dir, cou_data, rev_info_parsed, my_logger)


if __name__ == '__main__':
    main()
