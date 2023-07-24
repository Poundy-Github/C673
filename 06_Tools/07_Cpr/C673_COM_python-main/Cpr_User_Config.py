import openpyxl
import os,shutil
import re

DATA_TYPE_COUNT = 7 # /*SignalPhyDataType*/ 信号数据类型 所在的行数
RTE_FUNC_COUNT = 17 #/*Op_WriteSignalPhyData*/ 信号读写的RTE接口函数 所在的行数
STRUCT_ELEM_NUMBER = 20 #s_Cpr_ReceiveSignalConfig_t 结构体元素的数量
INPUT_SIGNAL_COUNT = 128 #Cpr_ReceiveSignalConfigTable 结构体数组的大小

INPUT_CHANGE_SOURCE_CPR_LCFG = "./input/Cpr_Lcfg.c"  #Doesn't seem to work??

#input file 
INPUT_CPR_FILENAME = "./input/base_tab.txt"
INPUT_FUNC_FILENAME = "./input/base_FUN.txt"

#output file,this file will be used to code Cpr_Lcfg.c
OUTPUT_CPR_FILENAME = "./output/Cpr_Signal_ConfigTable.txt"
OUTPUT_FUNC_FILENAME = "./output/ExtRte_Interface_Out.txt"
OUTPUT_RANGE_VALUE_FILENAME = "./output/Cpr_Signal_ValueRange.txt"

#just is middle variables
FUNC_TEMP_WORDS = "FunctionName"
RTE_TEMP_WORDS = "Rte_Write_PP_Cpr_SignalName"
TYPE_TEMP_WORDS = "data_type"

pattern = re.compile(r',(.*?) ')
pattern_type = re.compile(r'Cpr_SigDataType_(.*?) ')
pattern_signame = r'ComConf_ComSignal_SG(\w+)_\d+R'
pattern_type_ext = r'Cpr_SigDataType_(\w+)\s*/\*SignalRawDataType\*/'
wb_CAN0 = openpyxl.load_workbook('./input/CAN0_M.xlsx')
ws_CAN0 = wb_CAN0.active
wb_CAN1 = openpyxl.load_workbook('./input/CAN1_M.xlsx')
ws_CAN1 = wb_CAN1.active
lines_to_append = []

#genarate tem base_tab.txt
def Temporary_File():

    #open init file 
    with open(os.path.join(os.getcwd(), INPUT_CPR_FILENAME), "r") as ffirst:
        lines = ffirst.readlines()
    
    #add prefix "Ext_" and generate Temporary_File.txt
    with open('Temporary_File.txt', 'w') as ftemp:
        for line in lines: 
            line = line.replace('Ext_', '')
            ftemp.write(line)

#
def ExtRte_Interface_Func():

    with open(os.path.join(os.getcwd(), "Temporary_File.txt"), "r") as fbasetab:
        basetab_lines = fbasetab.readlines()
        fbasetab.seek(0)
        basetab_str = fbasetab.read()
    with open(os.path.join(os.getcwd(), INPUT_FUNC_FILENAME), "r") as fbasefun:
        rawbasefun = fbasefun.read()

    with open(os.path.join(os.getcwd(),OUTPUT_FUNC_FILENAME), "w") as foutput,open(os.path.join(os.getcwd(),OUTPUT_CPR_FILENAME), "w") as foutput_cpr:
        for i in range(INPUT_SIGNAL_COUNT):
            if "Rte_Write_PP_Cpr" in basetab_lines[RTE_FUNC_COUNT + i*STRUCT_ELEM_NUMBER]:
                result = pattern.search(basetab_lines[RTE_FUNC_COUNT + i*STRUCT_ELEM_NUMBER])
                if result:
                    newonefunc = rawbasefun.replace(FUNC_TEMP_WORDS,"Ext_" + result.group(1))
                    newonefunc = newonefunc.replace(RTE_TEMP_WORDS,result.group(1))
                    result_type = pattern_type.search(basetab_lines[DATA_TYPE_COUNT + i*STRUCT_ELEM_NUMBER])
                    newonefunc = newonefunc.replace(TYPE_TEMP_WORDS,result_type.group(1).lower())
                    foutput.write(newonefunc + '\n') 
                else:
                    print("null date!")
        new_basetab_str = re.sub(r'Rte_Write_PP_Cpr_', 'Ext_Rte_Write_PP_Cpr_', basetab_str)
        foutput_cpr.write(new_basetab_str)
    print("Creat_func and out_cpr_tab.txt end!")
    return


def ExtRte_Declaration():
    #打开文件,并读取其中的每一行
    with open(OUTPUT_FUNC_FILENAME, 'r') as f:
        for line in f:
        # 对于每一行,在其中查找是否包含字符串 'void Ext_Rte_Write_PP_Cpr_
            if 'void Ext_Rte_Write_PP_Cpr_' in line:
                #如果包含该字符串,则将该行文本加入到一个列表中
                line = line.rstrip()  + ';\n'
                lines_to_append.append(line)
            
    #在读取完文件中的所有行之后,关闭文件,并打开同一个文件,以追加模式打开
    with open(OUTPUT_FUNC_FILENAME, 'a') as f:
        #将列表中的所有行写入到文件中
        for line in lines_to_append:
            f.write(line + '\n')
    print("Creat_func_declaration end!  Ext_Rte_Write counters:",len(lines_to_append))
    
    #os.remove("base_tab_temp.txt")
    return

#process signal raw range 
def Cpr_Rawvalue_Range_Process():
    counters = 0
    dict_name = {}
    with open(os.path.join(os.getcwd(), INPUT_CPR_FILENAME), "r") as fbasetab:
        basetab_str = fbasetab.read()
        signalname = re.findall(pattern_signame, basetab_str)
        signaltype = re.findall(pattern_type_ext, basetab_str)
    lens = len(signalname)
    for i in range(0,lens):
        dict_name[signalname[i]] = signaltype[i]
    #print(dict_name)

    with open(os.path.join(os.getcwd(), OUTPUT_RANGE_VALUE_FILENAME), "w") as f:
        for row in ws_CAN0.iter_rows(min_row=2, min_col=3, max_col=3):
            if row[0].value in dict_name:
                counters += 1

                tpye_1 = dict_name.get(row[0].value)
                # print(tpye_1)
                Signal_offset = ws_CAN0.cell(row=row[0].row, column=10).value 
                Signal_factor = ws_CAN0.cell(row=row[0].row, column=9).value
                CAN0_MIN_value = ws_CAN0.cell(row=row[0].row, column=11).value 
                CAN0_MAX_value = ws_CAN0.cell(row=row[0].row, column=12).value
                row_min = round((float(CAN0_MIN_value) - float(Signal_offset)) / float(Signal_factor))
                row_max = round((float(CAN0_MAX_value) - float(Signal_offset)) / float(Signal_factor))

                str1 = 'const ' + tpye_1.lower() +' Cpr_Min_'+ row[0].value + ' = ' + str(row_min) + ';' +'\n'
                str2 = 'const ' + tpye_1.lower() +' Cpr_Max_'+ row[0].value + ' = ' + str(row_max) + ';' +'\n'
                f.write(str1 + str2 + "\n")
                del dict_name[row[0].value]
    
    #print(dict_name)
    with open(os.path.join(os.getcwd(), OUTPUT_RANGE_VALUE_FILENAME), "a") as f:
        for row in ws_CAN1.iter_rows(min_row=2, min_col=3, max_col=3):
            if row[0].value in dict_name:
                counters += 1
                tpye_2 = dict_name.get(row[0].value)
                # print('type2:', tpye_2)
                Signal_offset = ws_CAN1.cell(row=row[0].row, column=11).value 
                Signal_factor = ws_CAN1.cell(row=row[0].row, column=10).value
                CAN1_MIN_value = ws_CAN1.cell(row=row[0].row, column=12).value 
                CAN1_MAX_value = ws_CAN1.cell(row=row[0].row, column=13).value
                row_min = round((float(CAN1_MIN_value) - float(Signal_offset)) / float(Signal_factor))
                row_max = round((float(CAN1_MAX_value) - float(Signal_offset)) / float(Signal_factor))

                str1 = 'const ' + tpye_2.lower() +' Cpr_Min_'+ row[0].value + ' = ' + str(row_min) + ';' +'\n'
                str2 = 'const ' + tpye_2.lower() +' Cpr_Max_'+ row[0].value + ' = ' + str(row_max) + ';' +'\n'
                f.write(str1 + str2 + "\n")
                del dict_name[row[0].value]

    # print(dict_name)
    
    return

#check files exist 
def Check_file(infile):
    for file in infile:
        if not os.path.exists(file):
            raise Exception(f"File '{file}' does not exist.")


if __name__ == '__main__':

    #check files path valid 
    Check_file([INPUT_CPR_FILENAME,INPUT_FUNC_FILENAME])  #check files 

    #Genarate middle file to func used 
    Temporary_File()

    #Generate ExtRte function code and declarations
    ExtRte_Interface_Func()
    ExtRte_Declaration() 

    #output signal max min raw value 
    Cpr_Rawvalue_Range_Process()      

    print("succesfully!!!!")



