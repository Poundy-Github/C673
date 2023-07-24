#######################################################
#MKS - Drop all components in Legacy MKS projects and share from new mainstream repository location
#
#:org:           Continental AG
#:author:        Saad Azam
#:version:       $Revision: 1.1 $
#:date:          $Date: 2021/12/13 17:52:37CET $
#######################################################


import os
import sys
from openpyxl import Workbook, load_workbook


from optparse import OptionParser
from datetime import datetime, time
from mks_config_tools import mks_obj, logging_obj, xml_obj, ConfigSharedSubProjets_obj

class ShareComponentsFromMainstream_obj(ConfigSharedSubProjets_obj, mks_obj, logging_obj, xml_obj):
    ###########
    # Description: Class for automated check-in members
    ###########
    LOG_INFO = "Info"
    LOG_ERROR = "Error"
    LOG_WARN = "Warning"
    LOG_DEBUG = "Debug"
    NORMAL = 'normal'
    DONT_SHARE = 0

    def __init__(self):
        mks_obj.__init__(self)
        logging_obj.__init__(self)
        xml_obj.__init__(self)
        ConfigSharedSubProjets_obj.__init__(self)
        self.Host = "ims-adas"
        self.Port = "7001"
        self.cpid = ""
        # Flag used for  dryrun of MKS commands e.g. ConfigureSharedSubproject
        self.dryrun = False

    def ShareComponentsFromMainstream(self, options):
        ###########
        # ShareComponentsFromMainstream(options)
        #
        # Description: the ShareComponentsFromMainstream main function
        #
        # Parameter:    options specified in the input of main(argu)
        #
        # Return:         0 if ok 1 if errors occur
        ###########
        self.Host = options.Host
        self.Port = options.Port
        self.CpId = options.CpId
        self.Workbook = options.Workbook
        self.TargetWorksheet =  options.TargetWorksheet
        self.BaseProjPath = options.BaseProjPath
        self.BaseCompPath = options.BaseCompPath
        self.DevPath = options.DevPath
        self.SandboxPath = options.SandboxPath
        self.cell_info_list = []
        for value in options.cell_info_list.split(','):
            self.cell_info_list.append(int(value))
        self.Logging = options.Logging
        self.LogDir = options.LogDir
        self.UseDatabase = False

        #Lists for debugging
        self.list_mks_drop_failure = []
        self.list_mks_add_failure = []
        self.list_sharedsubproj_todrop_not_existing = []
        self.list_parent_subproj_not_existing = []
        self.list_subproj_toshare_not_existing = []

        error = 0

        # ********************** Setup Logging ******************************#
        Log_File_Main = "Log_" + self.DevPath + ".txt"
        Log_File_Pre = "Log_" + self.DevPath + "_pre.txt"
        Log_File_Post = "Log_" + self.DevPath + "_post.txt"
        self.LogFileName = Log_File_Main
        self.LogType[self.LOG_INFO] = 1
        self.LogType[self.LOG_WARN] = 1

        #do not remove this is needed by self.CreateSharedProjectXML()
        self.Sandbox = self.SandboxPath

        # ******************* Validate the base project, sandbox Path, logging dir and change package *****************
        BaseProjPath_mks = self.BaseProjPath + 'project.pj'
        if self.GetProjectView(BaseProjPath_mks, None, self.DevPath) is None:
            self.Log(self.LOG_INFO, "The input Base project path %s does not exist, please check" % BaseProjPath_mks)
            error = 1
            return error

        error = self.ValidateSandboxPath(self.SandboxPath, 'dir')
        error += self.validate_logdir_chpkg_sboxdp()
        if error != 0:
            return error
        # ********************** Read the excel sheet with the info about shared subprojects *************************
        path_dict_error, path_dict = self.read_mainstream_to_legacy_share_xls(self.Workbook, self.TargetWorksheet, self.cell_info_list,
                                                             self.BaseProjPath, self.BaseCompPath)
        if path_dict_error != 0:
            error = path_dict
            return error

        # ********************** Get shared subproject Pre-config ********************#
        self.Log(self.LOG_INFO, "***************************************************************************")
        self.Log(self.LOG_INFO, "Start getting list of pre configured shared subprojects from Target Sandbox")
        result1, rescanned_paths1 = self.CreateSharedProjectXML(Log_File_Pre, False, True)
        self.Log(self.LOG_INFO, "Stop getting list of pre configured shared subprojects from Target Sandbox")
        self.Log(self.LOG_INFO, "**************************************************************************")

        error = self.shared_scan_result_eval(result1)
        if error != 0:
            return error

        # ********************** Perform drop and reshare based on the info from excel sheet ********************#
        error = self.reshare_from_excel_sheet(path_dict)

        #********************** Get shared subproject Post-config ********************#
        self.Log(self.LOG_INFO, "***************************************************************************")
        self.Log(self.LOG_INFO, "Start getting list of post configured shared subprojects from Target Sandbox")
        #self.LogType["Debug"] = 0
        result2, rescanned_paths2 = self.CreateSharedProjectXML(Log_File_Post, False, True)
        self.Log(self.LOG_INFO, "Stop getting list of post configured shared subprojects from Target Sandbox")
        self.Log(self.LOG_INFO, "**************************************************************************")

        temp_error = self.shared_scan_result_eval(result2)
        if temp_error != 0:
            error = 1

        return error

    def validate_logdir_chpkg_sboxdp(self):
        ###########
        # validate_logdir_chpkg_sboxdp(self)
        #
        # Description: verify if the logging dir exist, change package is valid and if the sbox belong to the same devpath as target input devpath
        #
        # Parameter:    tbd
        #
        # Return:         0 if exist else 1
        ###########
        # check output log directory
        error = 0
        if ((os.path.isdir(self.LogDir) is False) and (self.Logging is True)):
            self.Log(self.ERROR, "log directory does not exist e.g. -o c:\temp \n")
            error = 1
        else:
            logTypes = {}
            logTypes[self.LOG_INFO] = 1
            #            logTypes["Warn"]=2
            #            logTypes["Debug"]=3
            logfile = self.LogFileName
            if self.Logging is True:
                logfile = os.path.join(self.LogDir, self.LogFileName)
                print "logfile=%s" % (logfile)
                if (os.path.isfile(logfile) is True):
                    if (self.RenameFileAndAddDateTime(logfile) != 0):
                        self.Log(self.LOG_ERROR, "RenameFileAndAddDateTime fails")

                error = self.SetupLogging(logfile, "file", self.LogType)
            else:
                error = self.SetupLogging(logfile, "console", self.LogType)
        if error != 0:
            print "Error: Setup Logging failed due to problems with LogType Format"
            return error
        # ********************** Validate the changepackage ********************#
        error = self.IsChangePackageValid(self.CpId, self.Host, self.Port)
        if error != 0:
            self.Log(self.LOG_INFO, 'Changepackage verification: Failed\n')
        else:
            self.Log(self.LOG_INFO, 'Changepackage verification: Passed\n')

        # *********************** Validate Sandbox devpath against input project DevPath name ***************

        sandbox_devpath = self.GetCurrentDevPath(self.SandboxPath)
        if sandbox_devpath == '':
            sandbox_devpath = self.NORMAL
        if sandbox_devpath != self.DevPath:
            error = 1
            self.Log(self.LOG_INFO,
                     'Sandbox Devpath %s and Input project devpath %s do not match, please check\n' % (
                     sandbox_devpath, self.DevPath))
        return error

    def validate_subproject_paths(self, ProjectPath, NewComponentPath):
        ###########
        # validate_subproject_paths(ProjectPath, NewComponentPath)
        #
        # Description: Validates target project path from where share should be dropped and new component subproject path 
        #              from where it should be added, and returns a dict of flags based on outcome.
        #
        # Parameter:    -ProjectPath: input subproject path
        #               -NewComponentPath: new component subproject path to be added/shared in the project
        #
        # Return:       flag_dict a dictionay with 3 flag enteries
        ###########
        parent_ProjectPath = ProjectPath[:ProjectPath.rfind('/', 0, ProjectPath.rfind('/'))]
        parent_ProjectPath = parent_ProjectPath + '/project.pj'

        flag_parent_projpath_exist = self.GetProjectView(parent_ProjectPath, None, self.DevPath)
        flag_no_share = self.does_share_exist_on_devpath(ProjectPath, self.DevPath)

        if NewComponentPath != '':
            flag_comp_loc_exist = self.GetProjectView(NewComponentPath, None, None)
            if flag_comp_loc_exist is None:
                self.list_subproj_toshare_not_existing.append(NewComponentPath)
                self.Log(self.LOG_INFO,
                         'New component location %s is invalid, please check if it exists' % NewComponentPath)
        else:
            flag_comp_loc_exist = None

        flag_dict = {'flag_parent_projpath_exist': flag_parent_projpath_exist, 'flag_no_share': flag_no_share,
                     'flag_comp_loc_exist': flag_comp_loc_exist}
        return flag_dict
    
    def does_share_exist_on_devpath(self, ProjectPath, DevPath):
        ###########
        # does_share_exist_on_devpath(ProjectPath,DevPath)
        #
        # Description: Verify if the input subproject path is valid for the given devpath
        #
        # Parameter:    -ProjectPath input subproject path
        #               -DevPath target development path
        #
        # Return:         0 if exist else 1
        ###########
        temp_ProjectPath = ProjectPath[:ProjectPath.rfind('/',0,ProjectPath.rfind('/'))]
        target_share = ProjectPath[ProjectPath.rfind('/',0,ProjectPath.rfind('/')) + 1:]
        temp_ProjectPath = temp_ProjectPath + '/project.pj'
        if DevPath == self.NORMAL:
            project_info = self.GetProjectView(temp_ProjectPath, None, None)
        else:
            project_info = self.GetProjectView(temp_ProjectPath, None, DevPath)

        flag_share = False
        if project_info is not None:
            for subproj in project_info:
                if target_share in subproj:
                    self.Log(self.LOG_INFO, 'Shared subproject %s exists in %s' % (ProjectPath,DevPath))
                    flag_share = True
                    break

        if flag_share is False:
            self.Log(self.LOG_INFO, 'Shared subproject %s doesn\'t exist in %s' % (ProjectPath, DevPath))
            return 1
        return 0

    def read_mainstream_to_legacy_share_xls(self, Workbook, TargetWorksheet, cell_info_list,base_project_path, base_old_comp_rep):
        ###########
        # read_mainstream_to_legacy_share_xls(Workbook, TargetWorksheet, cell_info_list,base_project_path, base_old_comp_rep)
        #
        # Description: Read specially formatted Excel sheet with project and component subproject locations to be
        # dropped and shared
        #
        # Parameter:    - Workbook: Input Excel sheet (http://ims-adas:7001/si/viewrevision?projectName=/ADAS/SW/Integr
        #                           ation/03%5fDocuments/Sensorics/project.pj&selection=Components%5fshareback%5flegacy.xlsx)
        #               - TargetWorksheet: Target worksheet to be read
        #               - cell_info_list: A comma seperated list with 5 integers e.g [2,3,4,8,78], 1st 3 integers
        #                                 represent the columns and last 2 represent start and ending row
        #               - base_project_path: Base project path e.g. /nfs/projekte1/PROJECTS/ARS400/06_Algorithm/
        #               - base_old_comp_rep: Base repository location of the components e.g. /nfs/projekte1/REPOSITORY/Base_Development/05_Algorithm/
        #
        # Return:         0 if the error occurs, else Dictionary with the {key:[value,value],..} (Project path :[old shared subproject path, new
        #                 shared subproject path])
        ###########
        proj_path_col = cell_info_list[0]
        old_comp_rep_col = cell_info_list[1]
        new_comp_rep_col = cell_info_list[2]
        row = cell_info_list[3]
        last_row = cell_info_list[4]

        notfound = True

        error = 0
        dict = {}
        # ************************** validate the cell_info_list ******************************************
        DEFAULT_PROJECT_PATH_COL = 5
        DEFAULT_OLD_COMP_COL = 6
        DEFAULT_NEW_COMP_COL = 7
        DEFAULT_STARTING_ROW = 10

        if proj_path_col != DEFAULT_PROJECT_PATH_COL or \
                        old_comp_rep_col != DEFAULT_OLD_COMP_COL or \
                        new_comp_rep_col != DEFAULT_NEW_COMP_COL:
            self.Log(self.LOG_WARN, "Input Excel worksheet Formatting problems, its recommended to have project path in column 5(E),"
                                     " old component repo in column 6(F) and new component repo in column 7(G)\n")

        if row != DEFAULT_STARTING_ROW:
            self.Log(self.LOG_WARN, "Input Excel worksheet Formatting problem, are you sure you want to set starting row != 10\n")

        input = raw_input("Is correct --CellInfoList provided i.e. %s (proj_path_col, old_comp_rep_col, new_comp_rep_col" \
                          ", first_row, last_row)\nEnter y for yes and n for no: " % cell_info_list)
        if input == 'n':
            error = 1
            return error, dict
        elif input != 'y':
            error = 1
            self.Log(self.LOG_ERROR, "Wrong input please use y or n, please retry, exiting...")
            return error, dict
        # *************************************************************************************************
        workbook_exist = os.path.isfile(self.Workbook)
        if not workbook_exist:
            error = 1
            self.Log(self.LOG_ERROR, "Input Workbook path is invalid, %s please make sure it exists" % self.Workbook)
            return error, dict
        else:
            Workbook = load_workbook(Workbook)
            for worksheet in Workbook:
                if worksheet.title == TargetWorksheet:
                    notfound = False
                    self.Log(self.LOG_INFO, "Input Worksheet title: %s" % (worksheet.title))
                    for count in range(last_row-row + 1):
                        base_path = base_project_path + worksheet.cell(row=row, column=proj_path_col).value

                        rel_old_loc = worksheet.cell(row=row, column=old_comp_rep_col).value
                        if rel_old_loc is not None:
                            old_comp_path = base_old_comp_rep + worksheet.cell(row=row, column=old_comp_rep_col).value
                        else:
                            old_comp_path = ''

                        new_comp_path = worksheet.cell(row=row, column=new_comp_rep_col).value
                        if new_comp_path is None:
                            new_comp_path = ''

                        dict[base_path] = [old_comp_path, new_comp_path]
                        row += 1
            if notfound is True:
                error = 1
                self.Log(self.LOG_ERROR, 'Input Worksheet %s does not seem to exist, please check' % TargetWorksheet)
        return error, dict

    def status_log (self,list_info,titel,logtype):
        ###########
        # status_log(list_info,titel,logtype)
        #
        # Description: Print out the status list for adding and dropping shares
        #
        # Parameter:   list_info: be one of the lists info about shared subprojects successfully or unsuccessfully added/dropped
        #              titel: Appropriate titel string 
        #              logtype: String for logging function logtype i.e "Error","Info","debug" etc
        #
        # Return:         0 if ok 1 if errors occur
        error = 0
        if len(list_info) > 0:
            error = 1
            self.Log(logtype, " " + "-"*len(titel))
            self.Log(logtype, titel)
            self.Log(logtype, " " + "-"*len(titel))
            for mem in list_info:
                self.Log(logtype, " " + mem + "\n")
        return error

    def shared_scan_result_eval(self,result):
        ###########
        # shared_scan_result_eval(result)
        #
        # Description: Evaluate and log if the shared subproject scanning in the project sandbox was ok
        #
        # Parameter:    - result: should be 0 or Non-Zero based on output of shared subproject parser
        #
        # Return:         0 if ok 1 if errors occur
        if result != 0:
            self.Log(self.LOG_ERROR, "Errors happened while getting pre-config share history")
            input = raw_input("Do you want to continue, Please enter y for yes and n for no: ")
            if input == 'n':
                return 1
            elif input != 'y':
                self.Log(self.LOG_ERROR, "Wrong input please use y or n, please retry, exiting...")
                return 1
        return 0

    def reshare_from_excel_sheet(self, path_dict):
        ###########
        # reshare_from_excel_sheet(argu)
        #
        # Description: Read info subproject path info from path dict and drop shared subprojects shared from old
        # component repository and re-share from Mainstream component repository
        #
        # Parameter:    - path_dict: should come from excel sheet as output of read_mainstream_to_legacy_share_xls()
        #                 or atleast the same format {target_proj_path:[old_comp_repo,new_comp_repo],....}
        # Return:         0 if ok 1 if errors occur
        ###########
        # *********************** Start reading excel sheet and drop/add share operation ********************
        for targetpath in path_dict:
            self.Log(self.LOG_INFO, '\n\n\n\n********************** Read from input Workbook **********************')
            self.Log(self.LOG_INFO, '%s \n %s \n %s' % (targetpath, path_dict[targetpath][0], path_dict[targetpath][1]))
            ProjectPath = targetpath
            OldComponentPath = path_dict[targetpath][0]
            NewComponentPath = path_dict[targetpath][1]

            self.Log(self.LOG_INFO,
                     '\n********************** Validate all paths in the project and component repository **************')

            flag_dict = self.validate_subproject_paths(ProjectPath, NewComponentPath)

            # Flag to test if parent subproject in the project where shared subproject must be dropped and reshared
            # exist -  Not None if exists else None
            flag_parent_projpath_exist = flag_dict['flag_parent_projpath_exist']
            # Flag to test if shared component subproject(to be dropped) exists in the project -  0 if exists else 1
            flag_no_share = flag_dict['flag_no_share']
            # Flag to testif New location of component subproject to be shared from exist - Not None if exists else None
            flag_comp_loc_exist = flag_dict['flag_comp_loc_exist']

            if flag_parent_projpath_exist is None:
                self.Log(self.LOG_INFO,
                         'Parent subproject of %s is invalid, please check if it exists' % ProjectPath)
                self.list_parent_subproj_not_existing.append(ProjectPath)
            else:
                self.Log(self.LOG_INFO, '\n**********************# Drop shared subprojects ***********************')
                if flag_no_share != self.DONT_SHARE:
                    self.Log(self.LOG_INFO,
                             'Cannot drop, the shared subproject, %s doesn\'t exist on %s' % (ProjectPath, self.DevPath))
                    self.list_sharedsubproj_todrop_not_existing.append(ProjectPath)
                else:
                    flag_drop_status = self.DropSubProject(ProjectPath, self.DevPath)
                    self.Log(self.LOG_INFO, 'Drop shared subproject error: %s' % flag_drop_status)
                    if flag_drop_status != 0:
                        self.Log(self.LOG_INFO,
                                 'Failed to drop the shared subproject %s on the Devpath: %s' % (ProjectPath, self.DevPath))
                        self.list_mks_drop_failure.append(ProjectPath)

                self.Log(self.LOG_INFO, '\n**********************# Add shared subprojects ************************\n')
                if flag_comp_loc_exist is not None:
                    if self.DevPath == self.NORMAL:
                        devpath = ''
                        flag_add_status = self.AddSharedSubProject(NewComponentPath, ProjectPath, devpath,
                                                                   SubprojectDevelopmentPath=None, SubprojectRevision='1.1',
                                                                   TypeOfShare="build")
                        self.Log(self.LOG_INFO, 'Add shared subproject error: %s' % flag_add_status)
                    else:
                        flag_add_status = self.AddSharedSubProject(NewComponentPath, ProjectPath, self.DevPath,
                                                                   SubprojectDevelopmentPath=None, SubprojectRevision='1.1',
                                                                   TypeOfShare="build")
                        self.Log(self.LOG_INFO, 'Add shared subproject error: %s' % flag_add_status)

                    if flag_add_status != 0:
                        self.list_mks_add_failure.append(ProjectPath)
                        self.Log(self.LOG_INFO,
                                 'Failed to add the shared subproject %s at the project location %s on the Devpath: %s' % (
                                 NewComponentPath, ProjectPath, self.DevPath))

        self.Log(self.LOG_INFO, "\n\n\n")
        Error_flag = 0
        Error_flag = self.status_log(self.list_parent_subproj_not_existing,
                                     'Following parent subprojects where the share must be added or dropped do not exist',
                                     self.LOG_ERROR)
        Error_flag += self.status_log(self.list_mks_drop_failure, 'Following shared subprojects failed to be dropped',
                                      self.LOG_ERROR)
        Error_flag += self.status_log(self.list_mks_add_failure, 'Following shared subprojects failed to be added',
                                      self.LOG_ERROR)
        Info1 = self.status_log(self.list_sharedsubproj_todrop_not_existing,
                                ' Following to-be-dropped shared subprojects do not exist', self.LOG_INFO)
        Error_flag += self.status_log(self.list_subproj_toshare_not_existing,
                                      'Following to-be-shared subprojects do not exist', self.LOG_ERROR)

        if Error_flag != 0:
            return 1

def main(argv):
    ###########
    # main()
    #
    # Description:  The main function to drop existing subprojects in legacy projects shared from old component
    #               repositories and re share from Mainstream component repository
    #
    # Author:       Saad Azam
    #
    # Parameters:   Command line arguments, see possible options and their descriptions below
    #
    # return 0 if successful otherwise return 1
    ###########
    usage = '%prog [options] Example:' \
            'python mks_reshare_mainstream_comp -H ims-adas -P 7001 -l -C 552986:3 --BaseProjPath /nfs/projekte1/PROJECTS/ARS400/06_Algorithm/' \
            '--SandboxPath D:\\mks\\AL_ARS410TA27_01.06.00_AI_Pilot\\project.pj --Workbook D:\\mks\\ADAS_SW\\Integration\\03_Documents\\Sensorics\\Components_shareback_legacy_test.xlsx' \
            '--TargetWorkSheet Components_Move_Decision_RADAR --DevPath AL_ARS410TA27_01.06.00_AI_Pilot --CellInfoList 5,6,7,10,14'
    parser = OptionParser(usage)
    parser.add_option("-H", "--Host", dest="Host",
                      help="e.g. mks-psad (mandatory)", default="ims-adas")
    parser.add_option("-P", "--Portconfig", dest="Port",
                      help=" e.g. 7001 (mandatory)", default="7001")
    parser.add_option("-C", "--ChangePackage", dest="CpId", default="",
                      help="changePackeageID e.g. 332131:1 (mandatory)")
    parser.add_option("-l", "--log", dest="Logging", action="store_true", default=False,
                      help="log(on/off)")
    parser.add_option("-o", "--output", dest="LogDir", default='.\\',
                      help="output log directory")
    parser.add_option("--BaseProjPath", dest="BaseProjPath", default='',
                      help="Base project path /nfs/projekte1/PROJECTS/ARS400/06_Algorithm/")
    parser.add_option("--BaseCompPath", dest="BaseCompPath",
                      default='/nfs/projekte1/REPOSITORY/Base_Development/05_Algorithm/',
                      help="Base project path /nfs/projekte1/REPOSITORY/Base_Development/05_Algorithm/")
    parser.add_option("", "--DevPath", dest="DevPath",
                      help="Target development paths, specify 'normal' for base project \
                      e.g normal,AL_ARS440DP26_02.02.00,AL_ARS4D1_04.02.00")
    parser.add_option("", "--SandboxPath", dest="SandboxPath",
                      help="Sandbox paths, of the project created from the same devpath as specified in --DevPath")
    parser.add_option("--Workbook", dest="Workbook", default='.\\',
                      help="Workbook i.e. excel document from where to read the info e.g /ADAS/SW/Integration/03_Documents/Sensorics/Components_shareback_legacy.xlsx")
    parser.add_option("", "--TargetWorkSheet", dest="TargetWorksheet",
                      help="Target Worksheet to read from in the Workbook")
    parser.add_option("", "--CellInfoList", dest="cell_info_list",
                      help="A comma seperated list with info on cells in the excel sheet to read info from i.e. proj_path_col, old_comp_rep_col, new_comp_rep_col,\
                      first_row, last_row e.g 5,6,7,10,71 ")

    (options, args) = parser.parse_args(argv)

    #call main class function to perform re-share
    error = ShareComponentsFromMainstream_obj().ShareComponentsFromMainstream(options)
    return error


if __name__ == '__main__':
    main(sys.argv)
