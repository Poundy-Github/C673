"""
Scripting to start a simulation run.

Features:
---------

Parses the following xmls and construct excel and text files.
        1. content of shared_project_config_rte.xml
    2. content of shared_project_config_master.xml (+ shared_project_config_interface.xml)
    3. content of shared_project_config_sim_bundle.xml
    4. content of shared_project_config_test_tools.xml
    5. content of shared_project_config_tools.xml

TODOs:
------


build_lib:
:org:           Continental AG
:author:        Lavanya Manickavasagam
:version:       $Revision: 1.18 $
:date:          $Date: 2017/10/24 12:39:29CEST $

"""
from getopt import getopt, GetoptError
from optparse import Option

MEMBER_VERSION = "$Revision: 1.18 $"
print "{0} {rev}".format(__file__, rev=MEMBER_VERSION)

import os.path

import sys, logging, os, datetime, win32com.client, collections, xml.etree.ElementTree as ET
import re

import openpyxl
from openpyxl.styles import Font, Style, PatternFill, Color, Border, Side, Alignment
from openpyxl.styles import colors


from collections import OrderedDict
import argparse


"""
getListofFiles(finalDict, configFileName, checkpoint_name, resultFileName, logger, ReleaseInfoDict)

Description: Get the list of xml files from the shared_project_config_files.xml
             Write in the text and excel files after parsing

Parameter: - finalDict - to store the complete details in dictionary format
        - configFileName -  the config file name from the command line argument
        - checkpoint_name - Checkpoint name from the command line argument
        - resultFileName - file name from the command line argument
"""
# python D:\My_reference\Pythonscripttask\generateModuleOverview.py -c D:\My_reference\Pythonscripttask\MFC430TA16_R03.05\shared_project_config_files.xml
#        -o D:\My_reference \ModuleOverview.xlsx  -cp  AL_MFC4T0_560A_03.05.00_INT23  -dp MFC4T0_560A_03.05 -ro  413243 -id 413241



def getListofFiles(finalDict, configFileName, checkpoint_name, resultDir, resultFileName, logger, ReleaseInfoDict,
                   hoCommentFile):
    try:
        logger = logging.getLogger("getListofFiles")
        logger.info('Entering..')

        now = datetime.datetime.now()
        rteDict = collections.OrderedDict()
        mast_inter_Dict = collections.OrderedDict()
        simBundleDict = collections.OrderedDict()
        testToolsDict = collections.OrderedDict()
        toolsDict = collections.OrderedDict()
        pathExists = os.path.isfile(configFileName)
        successFlag = True
        tree = None
        root = None
        xmlstring = None
        fileDirectory = None
        fileName = None
        xmlNotSuccess = None
        txtNotSuccess = None
        sequence = None
        global rteDictLength, mastInterDictLength, testToolsDictLength, toolsDictLength, simBundleDictLength

        if((not pathExists is None) and pathExists):
            logger.info('Given file exists ..')

            with open (configFileName, "r") as xmlfile:
                xmlstring = xmlfile.read().replace('\n', '')
                if xmlstring:
                    tree = ET.ElementTree(file=configFileName)
                    if tree is not None:
                        root = tree.getroot()
                        if root is not None:
                            fileDirectory = os.path.dirname(configFileName)
                            for projectFiles in root:
                                sequence = (fileDirectory, '\\', projectFiles.get('File'))
                                fileName = ''.join(sequence)
                                pathExists = os.path.isfile(fileName)
                                if((not pathExists is None) and pathExists):
                                    if "shared_project_config_rte.xml" in fileName:
                                        rteDict = parseXML(fileName, rteDict, logger)
                                        rteDictLength = len(rteDict)
                                    elif ("shared_project_config_master.xml" in fileName) or ("shared_project_config_base.xml" in fileName):
                                        mast_inter_Dict = parseXML(fileName, mast_inter_Dict, logger)
                                        mastInterDictLength = len(mast_inter_Dict)
                                    elif "shared_project_config_interface.xml" in fileName:
                                        mast_inter_Dict = parseXML(fileName, mast_inter_Dict, logger)
                                    elif "shared_project_config_test_tools.xml" in fileName:
                                        testToolsDict = parseXML(fileName, testToolsDict, logger)
                                        testToolsDictLength = len(testToolsDict)
                                    elif "shared_project_config_tools.xml" in fileName:
                                        toolsDict = parseXML(fileName, toolsDict, logger)
                                        toolsDictLength = len(toolsDict)
                                    elif "shared_project_config_sim_bundle.xml" in fileName:
                                        simBundleDict = parseXML(fileName, simBundleDict, logger)
                                        simBundleDictLength = len(simBundleDict)
                                else:
                                    logger.info('Please enter correct path of : ' + fileName)
                                    logger.info('Result of ' + fileName + ' is skipped')
                        else:
                            logger.info(configFileName + " is empty!!")
                            successFlag = False

                    else:
                        logger.info(configFileName + " is empty!!")
                        successFlag = False
        else:
            logger.info(' File : ' + configFileName + ' does not exist')

            successFlag = False


        finalDict = mergeAllDict(finalDict, logger, rteDict, mast_inter_Dict, testToolsDict, toolsDict, simBundleDict)
        if (not finalDict is None) and finalDict:
            logger.info('Componets exists..')
            if(os.path.isdir(resultDir)):
                xmlNotSuccess = writeXMLFile(resultDir, resultFileName, finalDict, checkpoint_name, logger,
                                             ReleaseInfoDict, hoCommentFile)
                txtNotSuccess = writeTextFile(resultDir, resultFileName, finalDict, logger)
            else:
                xmlNotSuccess = writeXMLFile(os.path.dirname(os.path.abspath(__file__)), resultFileName, finalDict,
                                             checkpoint_name, logger, ReleaseInfoDict, hoCommentFile)
                txtNotSuccess = writeTextFile(os.path.dirname(os.path.abspath(__file__)), resultFileName, finalDict,
                                              logger)


            if((xmlNotSuccess is None) or  xmlNotSuccess):
                logger.info('Result not written in xml..')
                # return 1
                successFlag = False
            if((txtNotSuccess is None) or  txtNotSuccess):
                logger.info('Result not written in txt..')
                logger.info('Program ended ..')
                # return 1
                successFlag = False
            logger.info('Finished writing results..')

        else:
            logger.info('No components available in the given xml. ')
            successFlag = False

    except IOError, err:
        logger.error('IO Error:')
        logger.error(err, exc_info=True)
        successFlag = False

    except IndentationError, err:
        logger.error('Indentation Error:')
        logger.error(err, exc_info=True)
        successFlag = False

    except Exception, err:
        logger.error('Error:')
        logger.error(err, exc_info=True)
        successFlag = False

    finally:
        logger.info('Exiting..')
        if(successFlag):
            return 0
        else:
            return 1

"""
writeTextFile(fileDirectory, resultFileName, finalDict, logger)

Description: Write result in the text file

Parameter: - fileDirectory - parent folder path where the txt file is to be saved
        - resultFileName -  Name of the text file
        - finalDict - dictionary to store the overall result

"""
def writeTextFile(fileDirectory, resultFileName, finalDict, logger):
    try:
        logger = logging.getLogger("writeTextFile")
        logger.info('writeTextFile Entering..')

        successFlag = True
        maxColLength = 45

        label1 = "Base Checkpoint Label"
        spaceToBeLeft1 = maxColLength - len(label1)
        spaceString1 = " "*spaceToBeLeft1

        label2 = "Custom Checkpoint Label"
        spaceToBeLeft2 = maxColLength - len(label2)
        spaceString2 = " "*spaceToBeLeft2

        label3 = "Interface Checkpoint Label"
        spaceToBeLeft3 = maxColLength - len(label3)
        spaceString3 = " "*spaceToBeLeft3

        maxLabelLength = None
        spaceToBeLeft = None
        spaceString = None


        sequence = (fileDirectory, '\\', resultFileName, '.txt')
        txtFilePath = ''.join(sequence)

        with open(txtFilePath, 'w') as fileObj:

            fileObj.write(label1 + spaceString1 + label2 + spaceString2 + label3 + spaceString3)
            fileObj.write("\n")

            for key, value in finalDict.iteritems():
                # maxLabelLenth - length of list of labelnames
                maxLabelLength = max(len(n) for n in value.values())
                if maxLabelLength is not None:
                    for i in range(0, maxLabelLength):
                        for key1, value1 in value.iteritems():
                            if 'BaseLabel' in key1:
                                if len(value1) > i:
                                    spaceToBeLeft = maxColLength - len(str(value1[i]))
                                    spaceString = " "*spaceToBeLeft
                                    fileObj.write(value1[i] + spaceString)
                                else:
                                    fileObj.write(" "*maxColLength)
                            elif 'CustomLabel' in key1:
                                if len(value1) > i:
                                    spaceToBeLeft = maxColLength - len(str(value1[i]))
                                    spaceString = " "*spaceToBeLeft
                                    fileObj.write(value1[i] + spaceString)
                                else:
                                    fileObj.write(" "*maxColLength)
                            elif 'InterfaceLabel' in key1:
                                if len(value1) > i:
                                    spaceToBeLeft = maxColLength - len(str(value1[i]))
                                    spaceString = " "*spaceToBeLeft
                                    fileObj.write(value1[i] + spaceString)
                                    fileObj.write("\n")
                                else:
                                    fileObj.write(" "*maxColLength)
                                    fileObj.write("\n")



    except IOError, err:
        logger.error('writeTextFile : IO Error:')
        logger.error(err, exc_info=True)
        logger.error('writeTextFile: Result not written in txt..')
        successFlag = False

    except IndentationError, err:
        logger.error('writeTextFile Indentation Error:')
        logger.error(err, exc_info=True)
        logger.error('writeTextFile : Result not written in txt..')
        successFlag = False

    except Exception, err:
        logger.error('writeTextFile Error:')
        logger.error(err, exc_info=True)
        logger.error('Result not written in txt..')
        successFlag = False

    finally:
        logger.info('writeTextFile Exiting..')
        if(successFlag):
            return 0
        else:
            return 1

"""
writeXMLFile(fileDirectory, resultFileName, finalDict, checkpoint_name, logger, ReleaseInfoDict)

Description: Write result in the xml file

Parameter: - fileDirectory - parent folder path where the txt file is to be saved
        - resultFileName -  Name of the text file
        - finalDict - dictionary to store the overall result
        -checkpoint_name - from the command line argument.  This is used to extract the sheet name under which the results are saved in the xml

"""
def writeXMLFile(fileDirectory, resultFileName, finalDict, checkpoint_name, logger, ReleaseInfoDict, hoCommentFile):

    try:

        logger = logging.getLogger("writeXMLFile")

        resultFileName_complete = os.path.join(fileDirectory, resultFileName + '.xlsx')
        logger.info('resultFileName_complete=%s' % (resultFileName_complete))
        # resultFileName_complete = fileDirectory+"\\"+resultFileName+".xlsx"
        successFlag = True
        workbook = None
        sheet1 = None
        rb = None
        wb = None
        row = 11  # 7th row at which Table starts
        rowInit = 1
        col = 1
        col1 = 1
        ColStrng = 'A'
        LabelStrings = [ '', 'Component Name', 'Base Checkpoint Label', 'Custom Checkpoint Label', 'Interface Checkpoint Label' ]
        LegendText = ['Color Legend Information:', 'Modified Component', 'Unmodified Component']
        ColumnWidth1 = 35
        ColumnWidth = 50
        RowHeight = 50
        NoColms = 5
        worksheet = None
        first_sheet_name = None


        Aerial14Font = Font(size=14, name='Arial')
        Aerial10Font = Font(size=10, name='Arial')
        styleObj = Style(font=Aerial14Font)
        LegendTextAlignment = Style(alignment=Alignment(horizontal='justify',
                                    vertical='justify',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0), font=Aerial10Font)
        TextAlignment = Style(alignment=Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0), font=Aerial14Font)

        CellColor = PatternFill(start_color='FFADFF2F',
                    end_color='FFADFF2F',
                    fill_type='solid')
        CellColorCompLst = PatternFill(start_color='FFDDDDDD',
                    end_color='FFDDDDDD',
                    fill_type='solid')
        CellColorBsCstIntLst = PatternFill(start_color='FFF2F2F2',
                    end_color='FFF2F2F2',
                    fill_type='solid')

        CellColorRTEComp = PatternFill(start_color='FFABC7EF',
                    end_color='FFABC7EF',
                    fill_type='solid')
        CellColorAlgoComp = PatternFill(start_color='FFFA6666',
                    end_color='FFFA6666',
                    fill_type='solid')
        CellColorToolComp = PatternFill(start_color='FF42CCEC',
                    end_color='FF42CCEC',
                    fill_type='solid')
        CellColorSimComp = PatternFill(start_color='FF93F5D7',
                    end_color='FF93F5D7',
                    fill_type='solid')

        CellColorDiff = PatternFill(start_color='FFFFFF0D',
                   end_color='FFFFFF0D',
                   fill_type='solid')

        ThinBorder = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

        ThickBorder = Border(left=Side(style='thick'),
                     right=Side(style='thick'),
                     top=Side(style='thick'),
                     bottom=Side(style='thick'))


        pathExists = os.path.isfile(resultFileName_complete)  # SM - Check if File exists on given Path

        if not pathExists:
            workbook = openpyxl.Workbook()  # SM - Create new Workbook ,add Empty Sheet and save the Workbook
            sheet1 = workbook.get_active_sheet()
            sheet1.title = 'Empty Sheet'
            workbook.save(resultFileName_complete)

        rb = openpyxl.load_workbook(resultFileName_complete)  # SM - Open existing Workbook
        # sheet = rb.get_sheet_by_name('checkpoint_name')
        # rb.save('resultFileName_complete')

        #Extract sheet name from checkpoint_name, sheet name should be shorter then check_point name due to
        #limited character length allowed in MS-Excel for sheet name.
        # Following is omitted from checkpoint_name:
        #Starting "AL_", characters like "." and "-"
        if checkpoint_name.find('AL_') != -1:
            sheetName = checkpoint_name[len('AL_'):]
            sheetName = re.sub('[.-]', '', sheetName)

        if (rb is not None) and (rb.get_sheet_names() is not None) :
            if (rb is not None)  :

                if sheetName in rb.get_sheet_names():
                    logger.info("Sheet Name '" + sheetName + "' already exists")
                    logger.info("Replacing existing Worksheet!!")
                    # logger.info('Result not written in xml..')
                    # Delete present worksheet
                    rb.remove_sheet(rb.get_sheet_by_name(sheetName))
                    # Create Blank Worksheet with same Name
                    worksheet = rb.create_sheet(index=0, title=sheetName)

                else:
                    strlen = len(sheetName)
                    if (strlen < 31) :
                        worksheet = rb.create_sheet(index=0, title=sheetName)  # SM - Add Work Sheet in Workbook
                    else:
                        logger.error("Error: Label string '" + sheetName + "' length exceeds 31 characters limit.")
                        successFlag = False


                if(successFlag and (not (worksheet is None))):

                    # Column 2 and 3 label is Check point label name
                    LabelRow = row - 6
                    ColumnCnt = col1 + 1
                    worksheet.cell(row=LabelRow, column=ColumnCnt).value = 'Algo Checkpoint :'
                    worksheet.cell(row=LabelRow, column=ColumnCnt).style = TextAlignment
                    worksheet.cell(row=LabelRow, column=ColumnCnt).fill = CellColorCompLst  # At 5th Row 2nd Column inserting Checkpoint label
                    worksheet.cell(row=LabelRow, column=ColumnCnt).border = ThickBorder
                    worksheet.cell(row=LabelRow, column=ColumnCnt + 1).value = checkpoint_name
                    worksheet.cell(row=LabelRow, column=ColumnCnt + 1).style = TextAlignment
                    worksheet.cell(row=LabelRow, column=ColumnCnt + 1).fill = CellColorCompLst  # At 5th Row 3rd Column inserting Checkpoint label
                    worksheet.cell(row=LabelRow, column=ColumnCnt + 1).border = ThickBorder





                    LabelRow = row - 4
                    ColumnCnt = col1 + 1
                    for key in sorted(ReleaseInfoDict) :
                        temp = ReleaseInfoDict[key]
                        for items in range(len(temp)):
                            worksheet.cell(row=LabelRow, column=ColumnCnt).value = temp[items]
                            worksheet.cell(row=LabelRow, column=ColumnCnt).fill = CellColorCompLst
                            worksheet.cell(row=LabelRow, column=ColumnCnt).border = ThinBorder
                            ColumnCnt = ColumnCnt + 1

                        LabelRow = LabelRow + 1
                        ColumnCnt = col1 + 1


                    worksheet.row_dimensions[row].height = RowHeight
                    for colCnt in range(NoColms):  # Set width(80), Font and Background colour  for 4 columns A1, B1, C1 and D1
                    # row = row
                        colCnt = colCnt + 1
                        worksheet.cell(row=row, column=colCnt).value = LabelStrings[colCnt - 1]
                        worksheet.cell(row=row, column=colCnt).style = TextAlignment
                        worksheet.cell(row=row, column=colCnt).fill = CellColor
                        worksheet.cell(row=row, column=colCnt).border = ThickBorder

                        if colCnt is 1:
                            worksheet.column_dimensions[ColStrng].width = ColumnWidth1  # Set Width of Column
                        else:
                            worksheet.column_dimensions[ColStrng].width = ColumnWidth

                        ColStrng = chr(ord('A') + colCnt)  # Get next Column Index


                    # merge cell from row 2 till number of RTE cell
                    worksheet.merge_cells(start_row=row + 1, start_column=1, end_row=rteDictLength + row, end_column=1)
                    worksheet.cell(row=row + 1, column=1).value = 'SW Common RTE'
                    worksheet.cell(row=row + 1, column=1).style = TextAlignment
                    worksheet.cell(row=row + 1, column=1).fill = CellColorRTEComp
                    worksheet.cell(row=row + 1, column=1).border = ThickBorder

                    NextStartRow = row + (rteDictLength + 1)
                    NextEndRow = mastInterDictLength + (NextStartRow - 1)

                    # merge (row 2 + rteDictLength ) till  mastInterDictLength cell
                    worksheet.cell(row=NextStartRow, column=1).value = 'Algo Components'
                    worksheet.cell(row=NextStartRow, column=1).style = TextAlignment
                    worksheet.cell(row=NextStartRow, column=1).fill = CellColorAlgoComp
                    worksheet.cell(row=NextStartRow, column=1).border = ThickBorder
                    worksheet.merge_cells(start_row=NextStartRow, start_column=1, end_row=NextEndRow, end_column=1)
                    AddToolLenght = testToolsDictLength + toolsDictLength
                    NextStartRow = row + (mastInterDictLength + rteDictLength + 1)
                    NextEndRow = AddToolLenght + (NextStartRow - 1)

                    worksheet.cell(row=NextStartRow, column=1).value = 'Tools / Test Tools'
                    worksheet.cell(row=NextStartRow, column=1).style = TextAlignment
                    worksheet.cell(row=NextStartRow, column=1).fill = CellColorToolComp
                    worksheet.cell(row=NextStartRow, column=1).border = ThickBorder
                    worksheet.merge_cells(start_row=NextStartRow, start_column=1, end_row=NextEndRow, end_column=1)
                    NextStartRow = row + (AddToolLenght + (mastInterDictLength + rteDictLength + 1))
                    NextEndRow = simBundleDictLength + (NextStartRow - 1)

                    worksheet.cell(row=NextStartRow, column=1).value = 'Sim Bundle'
                    worksheet.cell(row=NextStartRow, column=1).style = TextAlignment
                    worksheet.cell(row=NextStartRow, column=1).fill = CellColorSimComp
                    worksheet.cell(row=NextStartRow, column=1).border = ThickBorder
                    worksheet.merge_cells(start_row=NextStartRow, start_column=1, end_row=NextEndRow, end_column=1)

                    NextStartRow = NextEndRow + 2

                    # Get present  and previous worksheet index
                    worksheetPrev = rb.worksheets[1]
                    rb.save(resultFileName_complete)

                    for key, value in finalDict.iteritems():
                        row = row + 1
                        worksheet.cell(row=row, column=col + 1).value = key
                        worksheet.cell(row=row, column=col + 1).fill = CellColorCompLst
                        worksheet.cell(row=row, column=col + 1).border = ThinBorder
                        for key1, value1 in value.iteritems():

                            if 'BaseLabel' in key1:
                                ColCnt = col + 1
                                worksheet.cell(row=row, column=col + ColCnt).value = (", ".join(str(e) for e in value1))
                                worksheet.cell(row=row, column=col + ColCnt).fill = CellColorBsCstIntLst
                                worksheet.cell(row=row, column=col + ColCnt).border = ThinBorder

                            elif 'CustomLabel' in key1:
                                ColCnt = col + 2
                                worksheet.cell(row=row, column=col + ColCnt).value = (", ".join(str(e) for e in value1))
                                worksheet.cell(row=row, column=col + ColCnt).fill = CellColorBsCstIntLst
                                worksheet.cell(row=row, column=col + ColCnt).border = ThinBorder

                            elif 'InterfaceLabel' in key1:
                                ColCnt = col + 3
                                worksheet.cell(row=row, column=col + ColCnt).value = (", ".join(str(e) for e in value1))
                                worksheet.cell(row=row, column=col + ColCnt).fill = CellColorBsCstIntLst
                                worksheet.cell(row=row, column=col + ColCnt).border = ThinBorder

                            # Change the cell background colour if there are changes in cell values
                            current_val = str(worksheet.cell(row=row, column=col + ColCnt).value)
                            previous_val = str(worksheetPrev.cell(row=row, column=col + ColCnt).value)
                            if (previous_val != 'None'):  # Active sheet empty cell and previous worksheet empty cell contents are not same. This if loop to tackle this problem
                                if len(current_val) != 0 :  # Previous cell and present cell have different contents and current cell contains string
                                    if (current_val != previous_val) :
                                        worksheet.cell(row=row, column=col + ColCnt).fill = CellColorDiff
                                elif len(current_val) == 0 :  # Previous cell and present cell have different contents and current cell contains No string
                                    worksheet.cell(row=row, column=col + ColCnt).fill = CellColorDiff
                            elif len(current_val) != 0 :
                                worksheet.cell(row=row, column=col + ColCnt).fill = CellColorDiff


                    # Set Legend Information
                    row = 1
                    worksheet.cell(row=row, column=1).value = LegendText[0]
                    worksheet.cell(row=row, column=1).style = TextAlignment
                    worksheet.cell(row=row, column=1).border = ThickBorder

                    worksheet.cell(row=row + 1, column=1).value = LegendText[1]
                    worksheet.cell(row=row + 1, column=1).style = LegendTextAlignment
                    worksheet.cell(row=row + 1, column=1).fill = CellColorDiff
                    worksheet.cell(row=row + 1, column=1).border = ThickBorder

                    worksheet.cell(row=row + 2, column=1).value = LegendText[2]
                    worksheet.cell(row=row + 2, column=1).style = LegendTextAlignment
                    worksheet.cell(row=row + 2, column=1).fill = CellColorBsCstIntLst
                    worksheet.cell(row=row + 2, column=1).border = ThickBorder

                    writeHoCommentIntoExcelSheet(worksheet, NextStartRow, fileDirectory, hoCommentFile, TextAlignment)

                    rb.save(resultFileName_complete)
                    first_sheet_name = rb.get_sheet_names()[0]
                    if first_sheet_name is not None:
                        moveSheet(resultFileName_complete, checkpoint_name, first_sheet_name, logger)


                else:
                    logger.info('No worksheet in XML..')
                    successFlag = False

            else:
                logger.info('XML could not be copied..')
                successFlag = False
        else:
            logger.info('XML could not be opened..')
            successFlag = False



    except IOError, err:
        logger.error('IO Error:')
        if err.errno == 13:
            logger.error('Please close the result file : ' + resultFileName_complete)
        else:
            logger.error(err, exc_info=True)
        logger.error('Result not written in xml..')
        successFlag = False


    except IndentationError, err:
        logger.error('Indentation Error:')
        logger.error(err, exc_info=True)
        logger.error('Result not written in xml..')
        successFlag = False

    except ValueError, err:
        logger.error(err, exc_info=True)
        logger.error('Checkpoint label is more than 31 characters: ')
        logger.error('Result not written in xml..')
        successFlag = False


    except Exception, err:
        logger.error('Error:')
        logger.error(err, exc_info=True)
        logger.error('Result not written in xml..')
        successFlag = False

    finally:
        logger.info('writeXMLFile Exiting..')
        if(successFlag):
            return 0
        else:
            return 1


def writeHoCommentIntoExcelSheet(worksheet, NextStartRow, fileDirectory, hoCommentFile, TextAlignment):
    logger = logging.getLogger("writeHoCommentIntoExcelSheet")

    sequence = (fileDirectory, '\\', hoCommentFile)
    txtFilePath = ''.join(sequence)

    if os.path.exists(txtFilePath):

        logger.info("Going to write comment of file '" + txtFilePath + "' into excel sheet.")

        with open(txtFilePath, 'r') as f:

            text = 'Comments (please double-klick into this area to wrap text): \r\n'
            lineCounter = 0
            firstIter = True

            for line in f.read().split('\n'):
                if (firstIter == False):
                    text += '\r\n'

                text += line
                firstIter = False
                lineCounter += 1

        logger.info("Following text will be moved into Excel Sheet: '" + text + "'")

        worksheet.cell(row=NextStartRow, column=1).style = TextAlignment
        worksheet.cell(row=NextStartRow, column=1).value = text
        worksheet.cell(row=NextStartRow, column=5).alignment = Alignment(wrapText=True)
        worksheet.merge_cells(start_row=NextStartRow, start_column=1, end_row=NextStartRow + lineCounter, end_column=5)

    else:
        logger.info("No handover-comment-file given. Do nothing.")

"""
mergeAllDict(finalDict, logger, rteDict, mast_inter_Dict, testToolsDict, toolsDict, simBundleDict)

Description: Merge all the sub dictionaries - done for the sorting purspose

Parameter: - finalDict - master dictionary containing all the sub dictionaries
           - rteDict - results from shared_project_config_rte.xml
           - mast_inter_Dict - results from shared_project_config_master.xml and shared_project_config_interface.xml
           - testToolsDict - results from shared_project_config_test_tools.xml
           - toolsDict - results from shared_project_config_tools.xml
           - simBundleDict - results from shared_project_config_sim_bundle.xml

"""
def mergeAllDict(finalDict, logger, rteDict, mast_inter_Dict, testToolsDict, toolsDict, simBundleDict):
    logger = logging.getLogger("mergeAllDict")
    logger.info('Entering..')
    successFlag = True
    try:
        finalDict = rteDict.copy()
        finalDict.update(mast_inter_Dict)
        finalDict.update(testToolsDict)
        finalDict.update(toolsDict)
        finalDict.update(simBundleDict)

    except Exception, err:
        logger.error('Error:')
        logger.error(err, exc_info=True)
        successFlag = False
    finally:
        logger.info('Exiting..')
        if(successFlag):
            return finalDict
        else:
            return None



"""
setFileDict(localFileName,subDict, logger)

Description: Based on the filename assign the dictionary

Parameter: - FileName - Name of the xml being parsed
           - subDict - dictionary to be stored
"""
# #def setFileDict(localFileName,subDict, logger):
# #    if "shared_project_config_rte.xml" in localFileName:
# #    rteDict = subDict
# #    elif "shared_project_config_master.xml" in localFileName:
# #    mast_inter_Dict = subDict
# #    elif "shared_project_config_interface.xml" in localFileName:
# #    mast_inter_Dict = subDict
# #    elif "shared_project_config_test_tools.xml" in localFileName:
# #    testToolsDict = subDict
# #    elif "shared_project_config_tools.xml" in localFileName:
# #    toolsDict = subDict
# #    elif "shared_project_config_sim_bundle.xml" in localFileName:
# #    simBundleDict = subDict

"""
moveSheet(fileName, sheetName, first_sheet_name, logger)

Description: Move the latest sheet to the first position

Parameter: fileName - Complete path of the xlsx file,
           sheetName - Name of the sheet to be moved to the first index position
           first_sheet_name - name of the first sheet present in the already existing xlsx
"""
def moveSheet(fileName, sheetName, first_sheet_name, logger):
    logger = logging.getLogger("moveSheet")
    logger.info('Entering..')
    successFlag = True
    excel = None
    wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        if not (excel is None):
            wb = excel.Workbooks.Open(Filename=fileName, ReadOnly='False')
            if not (wb is None):
                if not (wb.Sheets is None):
                    for worksheet in wb.Sheets:
                        if worksheet.Name == sheetName:
                            worksheet.Move(Before=wb.Sheets(first_sheet_name))
                    wb.Save()
                    wb.Close()
            else:
                successFlag = False
                logger.info('workbook could not be opened..')
        else:
            successFlag = False
            logger.info('No excel object found..')
    # return 0

    except IOError, err:
        logger.error('IO Error:')
        logger.error(err, exc_info=True)
        successFlag = False

    except IndentationError, err:
        logger.error('Indentation Error:')
        logger.error(err, exc_info=True)
        successFlag = False

    except Exception, err:
        logger.error('Error:')
        logger.error(err, exc_info=True)
        successFlag = False

    finally:
        logger.info('Exiting..')
        if(successFlag):
            return 0
        else:
            return 1


"""
getFileNumber(localFileName, logger)

Description: Add numbers for the reference purpose

Parameter: -FileName - Name of the xml being parsed
"""
# #def getFileNumber(localFileName, logger):
# #    try:
# #    if "shared_project_config_rte.xml" in localFileName:
# #        return 1
# #    elif "shared_project_config_master.xml" in localFileName:
# #        return 2
# #    elif "shared_project_config_interface.xml" in localFileName:
# #        return 3
# #        elif "shared_project_config_test_tools.xml" in localFileName:
# #        return 4
# #    elif "shared_project_config_tools.xml" in localFileName:
# #        return 5
# #    elif "shared_project_config_sim_bundle.xml" in localFileName:
# #        return 6
# #    except Exception,err:
# #        logger.error('Error:')
# #    logger.error(err, exc_info=True)
# #    return 0

"""
getFileDict(localFileName, logger)

Description: Get the specific dictionary based on the xml file being parsed

Parameter: - localFileName - Name of the xml file being parsed

"""
# #def getFileDict(localFileName, logger):
# #    try:
# #    if "shared_project_config_rte.xml" in localFileName:
# #        return rteDict
# #    elif "shared_project_config_master.xml" in localFileName:
# #        return mast_inter_Dict
# #    elif "shared_project_config_interface.xml" in localFileName:
# #        return mast_inter_Dict
# #    elif "shared_project_config_sim_bundle.xml" in localFileName:
# #        return simBundleDict
# #    elif "shared_project_config_test_tools.xml" in localFileName:
# #        return testToolsDict
# #    elif "shared_project_config_tools.xml" in localFileName:
# #        return toolsDict
# #    except Exception,err:
# #    logger.error('Error:')
# #    logger.error(err, exc_info=True)
# #    return 0

"""
parseXML(fileName, tempDict, logger)

Description: parses the given xml and constructs the dictionary

Parameter: - fileName - FileName of the xml to be parsed
           - tempDict - Dictionary where the result is stored
"""
def parseXML(fileName, tempDict, logger):

    # localFileName, tree, root, baseConfigList, sharedProject, config, labelName, compName, localDict, localValDict, customConfigList, customSharedProject, customConfig, interSharedProject, interConfig
    logger = logging.getLogger("parseXML")
    logger.info('Entering..')
    logger.info("Parsing : " + fileName)
    localFileName = os.path.basename(fileName)
    successFlag = True
    tree = None
    root = None
    baseConfigList = None
    labelName = None
    compName = None
    localDict = None
    localValDict = None
    customConfigList = None
    interConfigList = None

    try:
        if not 'interface' in fileName:
            tree = ET.parse(fileName)
            if tree is not None:
                root = tree.getroot()
                if root is not None:
                    baseConfigList = root.find('BaseConfig')
                    if baseConfigList is not None:
                        for sharedProject in root.find('BaseConfig'):
                            for config in sharedProject:
                                labelName = config.get('Label')
                                ConfigType = config.get('Type')
                                compName = sharedProject.get('CompName')
                                if ConfigType == 'build':
                                    localDict = collections.OrderedDict()
                                    if compName in tempDict:
                                        localDict = tempDict.get(compName)
                                        localDict['BaseLabel'].append(labelName)
                                        tempDict[compName] = localDict
                                    else:
                                        localValDict = collections.OrderedDict()
                                        localValDict['BaseLabel'] = []
                                        localValDict['CustomLabel'] = []
                                        localValDict['InterfaceLabel'] = []
                                        localValDict['fileNumber'] = []
                                        localValDict['fileName'] = []
                                        localValDict['fileName'].append(localFileName)
                                        # localValDict['fileNumber'].append(getFileNumber(localFileName, logger))
                                        localValDict['BaseLabel'].append(labelName)
                                        tempDict.update({compName: localValDict})
                    customConfigList = root.find('CustomConfigV2')
                    if customConfigList is not None:
                        for customSharedProject in root.find('CustomConfigV2'):
                            for customConfig in customSharedProject:
                                labelName = customConfig.get('Label')
                                ConfigType = config.get('Type')
                                compName = customSharedProject.get('CompName')
                                # if ConfigType == 'build':
                                if labelName:
                                    localDict = collections.OrderedDict()
                                    if compName in tempDict:
                                        localDict = tempDict.get(compName)
                                        localDict['CustomLabel'].append(labelName)
                                        tempDict[compName] = localDict
                                    else:
                                        localValDict = collections.OrderedDict()
                                        localValDict['BaseLabel'] = []
                                        localValDict['CustomLabel'] = []
                                        localValDict['InterfaceLabel'] = []
                                        localValDict['fileNumber'] = []
                                        localValDict['fileName'] = []
                                        localValDict['fileName'].append(localFileName)
                                        # localValDict['fileNumber'].append(getFileNumber(localFileName, logger))
                                        localValDict['CustomLabel'].append(labelName)
                                        tempDict.update({compName: localValDict})


        if 'interface' in fileName:
            tree = ET.parse(fileName)
            if tree is not None:
                root = tree.getroot()
                if root is not None:
                    interConfigList = root.find('CustomConfigV2')
                    if  interConfigList is not None:
                        for interSharedProject in root.find('CustomConfigV2'):
                            for interConfig in interSharedProject:
                                labelName = interConfig.get('Label')
                                compName = interSharedProject.get('CompName')
                                if labelName:
                                    localDict = collections.OrderedDict()
                                    if compName in tempDict:
                                        localDict = collections.OrderedDict(tempDict.get(compName))
                                        localDict['InterfaceLabel'].append(labelName)
                                        tempDict[compName] = localDict
                                    else:
                                        localValDict = collections.OrderedDict()
                                        localValDict['BaseLabel'] = []
                                        localValDict['CustomLabel'] = []
                                        localValDict['InterfaceLabel'] = []
                                        localValDict['fileNumber'] = []
                                        localValDict['fileName'] = []
                                        localValDict['fileName'].append(localFileName)
                                        # localValDict['fileNumber'].append(getFileNumber(localFileName, logger))
                                        localValDict['InterfaceLabel'].append(labelName)
                                        tempDict.update({compName: localValDict})
        # return tempDict
    except IndentationError, err:
        logger.error('Indentation Error:')
        logger.error(err, exc_info=True)
        # return 1
        successFlag = False

    except Exception, err:
        logger.error('Error:')
        logger.error(err, exc_info=True)
        # return 1
        successFlag = False

    finally:
        logger.info('Exiting..')
        if(successFlag):
            return tempDict
        else:
            return {}


# Main Entry Point -------------------------------------------------------------
def main():


    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)
    logger.info('Main Function Logging Entering..')
    successFlag = False
    scriptSuccess = 1



    try:
        finalDict = collections.OrderedDict()
        outputFileName = ""
        opts = argparse.ArgumentParser(description="This script is used to read the configuration xmls and generates the report..")
        opts.add_argument("-c", "--config",
                      dest="config_file",
                      type=str,
                      help="Config file to be used. [mandatory]")
        opts.add_argument("-o", "--output",
                      dest="output_file",
                      type=str,
                      help="Output file to be used. [mandatory]")
        opts.add_argument("-cp", "--checkpoint",
                      dest="checkpoint_name",
                      type=str,
                      help="Checkpoint Label name. [mandatory]")
        opts.add_argument("-dp", "--intDevpath",
                      dest="Integration_development_path",
                      type=str,
                      help="Integration Development Path name. [mandatory]")
        opts.add_argument("-ro", "--rlsordr",
                      dest="Integration_ro",
                      type=str,
                      help="Integration RO. [mandatory]")
        opts.add_argument("-id", "--rlsID",
                      dest="IMS_release_id",
                      type=str,
                      help="IMS Release ID. [mandatory]")
        opts.add_argument("-ho", "--mocomment",
                      dest="ho_file",
                      type=str,
                      help="Input file which contain  handover_comments to be used. [optional]")
        args = opts.parse_args()

        checkpoint_name = args.checkpoint_name
        configFileName = args.config_file
        hoCommentFile = args.ho_file


        ReleaseInfoDict = {}
        ReleaseInfoDict = collections.OrderedDict()
        ReleaseInfoDict = {'ADevPathInt':("Integration Development Path", args.Integration_development_path),
                           'BIntRO':("Integration RO", args.Integration_ro), 'CIMSRelID':("IMS Release ID",
                                                                                          args.IMS_release_id)}

        if os.path.basename(args.output_file):
            outputFileName = os.path.splitext(os.path.basename(args.output_file))[0]
            outDir = os.path.dirname(args.output_file)


        if (configFileName is not None) and (checkpoint_name is not None) and (outputFileName is not None):
            if(configFileName and checkpoint_name and outputFileName):
            # if(configFileName and checkpoint_name and outputFileName):
                scriptSuccess = getListofFiles(finalDict, configFileName, checkpoint_name,
                                               outDir, outputFileName, logger, ReleaseInfoDict, hoCommentFile)
        if scriptSuccess is None:
            logger.error('Script execution NOT successful..')
            # return 1
            successFlag = False
        elif(scriptSuccess):
            logger.error('Script execution NOT successful1 ..')
            # return 1
            successFlag = False
        else:
            successFlag = True
            logger.info('Script execution successful..')
        # return 0


    except IndentationError, err:
        logger.error('Main Function Logging Indentation Error:')
        logger.error(err, exc_info=True)
        # return 1
        successFlag = False

    except Exception, err:
        logger.error('Main Function Logging Error:')
        logger.error(err, exc_info=True)
        # return 1
        successFlag = False

    finally:
        logger.info('Main Function Logging Exiting..')
        if(successFlag):
            return 0
        else:
            return 1

if (__name__ == "__main__"):
    res = main()
    print(res)
    sys.exit(res)
"""
CHANGE LOG:
-----------
$Log: gen_module_overview.py  $
Revision 1.18 2017/10/24 12:39:29CEST Arndt, Bosse (uidq5253) 
2017/1.047
Revision 1.16 2017/02/08 17:36:15CET Azam, Saad (uidg4889) 
Sheet name seperated from checkpoint label, although its still derived from it but is shorter.
Revision 1.14 2016/12/20 15:16:25CET Sinnwell-EXT, Wolfgang (uidj6607)
added -ho parameter
Revision 1.13 2016/09/29 16:14:43CEST Mali, Sangram (uidp6760)
Fixed Issue:
If there are empty entries in baseconfig. Script s not creating entry in Module Overview report.
Revision 1.12 2016/07/20 16:53:38CEST Sinnwell-EXT, Wolfgang (uidj6607)
added support for RADAR
Revision 1.11 2016/07/20 16:14:33CEST Mali, Sangram (uidp6760)
Added  Base Config Xml file for ARS
Revision 1.10 2016/04/21 12:44:16CEST Sinnwell-EXT, Wolfgang (uidj6607)
fixed missing sys.exit()
Revision 1.9 2016/03/17 13:32:37CET Sinnwell-EXT, Wolfgang (uidj6607)
added check for length of checkpoint_name
Revision 1.7 2016/02/02 11:29:52CET Sinnwell-EXT, Wolfgang (uidj6607) 
fixed typo "Label" and moved string "Algo Checkpoint" to new location
Revision 1.6 2016/01/27 11:48:25CET Azam, Saad (uidg4889) 
small format change
Revision 1.5 2016/01/25 16:51:59CET Sinnwell-EXT, Wolfgang (uidj6607) 
some more enhancements like highlighting changed label
Revision 1.3 2015/12/21 12:02:29CET Sinnwell-EXT, Wolfgang (uidj6607) 
fixed error handling on return

"""

